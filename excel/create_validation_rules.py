#!/usr/bin/env python3
"""
Create Excel Validation Rules
Generates comprehensive validation rules for the labor analytics system
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os


def create_validation_rules_excel():
    """Create comprehensive Excel validation rules file"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # =============================================
    # EMPLOYEES TABLE RULES
    # =============================================
    
    ws_employees = wb.create_sheet("Rules_employees")
    
    employee_rules = [
        ['employee_id_required', 'employee_id', 'required', True, 'Employee ID is required and cannot be null', True],
        ['employee_id_format', 'employee_id', 'regex', r'^EMP\d{3}$', 'Employee ID must be in format EMP001, EMP002, etc.', True],
        ['first_name_required', 'first_name', 'required', True, 'First name is required', True],
        ['last_name_required', 'last_name', 'required', True, 'Last name is required', True],
        ['email_format', 'email', 'regex', r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', 'Invalid email format', True],
        ['email_unique', 'email', 'unique', True, 'Email address must be unique', True],
        ['phone_format', 'phone', 'regex', r'^\+?1?\d{9,15}$', 'Phone number must be valid format', False],
        ['department_id_required', 'department_id', 'required', True, 'Department ID is required', True],
        ['department_id_exists', 'department_id', 'foreign_key', 'departments.department_id', 'Department ID must exist in departments table', True],
        ['position_required', 'position', 'required', True, 'Position is required', True],
        ['hire_date_required', 'hire_date', 'required', True, 'Hire date is required', True],
        ['hire_date_format', 'hire_date', 'data_type', 'date', 'Hire date must be a valid date', True],
        ['hire_date_not_future', 'hire_date', 'custom_formula', 'hire_date <= today()', 'Hire date cannot be in the future', True],
        ['termination_date_format', 'termination_date', 'data_type', 'date', 'Termination date must be a valid date', False],
        ['termination_after_hire', 'termination_date', 'custom_formula', 'termination_date > hire_date', 'Termination date must be after hire date', False],
        ['salary_required', 'salary', 'required', True, 'Salary is required', True],
        ['salary_range', 'salary', 'range', '20000:200000', 'Salary must be between $20,000 and $200,000', True],
        ['salary_positive', 'salary', 'range', '0:', 'Salary must be positive', True],
        ['hourly_rate_required', 'hourly_rate', 'required', True, 'Hourly rate is required', True],
        ['hourly_rate_range', 'hourly_rate', 'range', '10:200', 'Hourly rate must be between $10 and $200', True],
        ['hourly_rate_positive', 'hourly_rate', 'range', '0:', 'Hourly rate must be positive', True],
        ['employment_status_enum', 'employment_status', 'enum', 'ACTIVE,INACTIVE,TERMINATED,ON_LEAVE', 'Employment status must be one of: ACTIVE, INACTIVE, TERMINATED, ON_LEAVE', True],
        ['employment_status_default', 'employment_status', 'default', 'ACTIVE', 'Employment status defaults to ACTIVE', True]
    ]
    
    # Add headers
    headers = ['Rule Name', 'Field Name', 'Rule Type', 'Rule Value', 'Error Message', 'Is Active']
    for col, header in enumerate(headers, 1):
        cell = ws_employees.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data
    for row, rule in enumerate(employee_rules, 2):
        for col, value in enumerate(rule, 1):
            cell = ws_employees.cell(row=row, column=col, value=value)
            cell.border = border
    
    # Auto-adjust column widths
    for column in ws_employees.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_employees.column_dimensions[column_letter].width = adjusted_width
    
    # =============================================
    # TIME_TRACKING TABLE RULES
    # =============================================
    
    ws_time = wb.create_sheet("Rules_time_tracking")
    
    time_rules = [
        ['employee_id_required', 'employee_id', 'required', True, 'Employee ID is required', True],
        ['employee_id_exists', 'employee_id', 'foreign_key', 'employees.employee_id', 'Employee ID must exist in employees table', True],
        ['project_id_required', 'project_id', 'required', True, 'Project ID is required', True],
        ['project_id_exists', 'project_id', 'foreign_key', 'projects.project_id', 'Project ID must exist in projects table', True],
        ['date_required', 'date', 'required', True, 'Date is required', True],
        ['date_format', 'date', 'data_type', 'date', 'Date must be a valid date', True],
        ['date_not_future', 'date', 'custom_formula', 'date <= today()', 'Date cannot be in the future', True],
        ['hours_worked_required', 'hours_worked', 'required', True, 'Hours worked is required', True],
        ['hours_worked_range', 'hours_worked', 'range', '0:24', 'Hours worked must be between 0 and 24', True],
        ['hours_worked_positive', 'hours_worked', 'range', '0:', 'Hours worked must be positive', True],
        ['overtime_hours_range', 'overtime_hours', 'range', '0:16', 'Overtime hours must be between 0 and 16', True],
        ['overtime_hours_positive', 'overtime_hours', 'range', '0:', 'Overtime hours must be positive', True],
        ['overtime_calculation', 'overtime_hours', 'custom_formula', 'overtime_hours = max(0, hours_worked - 8)', 'Overtime hours should be calculated as max(0, hours_worked - 8)', True],
        ['task_description_length', 'task_description', 'length', '0:500', 'Task description must be 500 characters or less', False],
        ['billable_boolean', 'billable', 'data_type', 'boolean', 'Billable must be true or false', True],
        ['billable_default', 'billable', 'default', True, 'Billable defaults to true', True]
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws_time.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data
    for row, rule in enumerate(time_rules, 2):
        for col, value in enumerate(rule, 1):
            cell = ws_time.cell(row=row, column=col, value=value)
            cell.border = border
    
    # Auto-adjust column widths
    for column in ws_time.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_time.column_dimensions[column_letter].width = adjusted_width
    
    # =============================================
    # PAYROLL TABLE RULES
    # =============================================
    
    ws_payroll = wb.create_sheet("Rules_payroll")
    
    payroll_rules = [
        ['employee_id_required', 'employee_id', 'required', True, 'Employee ID is required', True],
        ['employee_id_exists', 'employee_id', 'foreign_key', 'employees.employee_id', 'Employee ID must exist in employees table', True],
        ['pay_period_start_required', 'pay_period_start', 'required', True, 'Pay period start date is required', True],
        ['pay_period_start_format', 'pay_period_start', 'data_type', 'date', 'Pay period start must be a valid date', True],
        ['pay_period_end_required', 'pay_period_end', 'required', True, 'Pay period end date is required', True],
        ['pay_period_end_format', 'pay_period_end', 'data_type', 'date', 'Pay period end must be a valid date', True],
        ['pay_period_end_after_start', 'pay_period_end', 'custom_formula', 'pay_period_end > pay_period_start', 'Pay period end must be after start', True],
        ['gross_pay_required', 'gross_pay', 'required', True, 'Gross pay is required', True],
        ['gross_pay_range', 'gross_pay', 'range', '0:50000', 'Gross pay must be between $0 and $50,000', True],
        ['gross_pay_positive', 'gross_pay', 'range', '0:', 'Gross pay must be positive', True],
        ['net_pay_required', 'net_pay', 'required', True, 'Net pay is required', True],
        ['net_pay_range', 'net_pay', 'range', '0:50000', 'Net pay must be between $0 and $50,000', True],
        ['net_pay_positive', 'net_pay', 'range', '0:', 'Net pay must be positive', True],
        ['net_pay_less_than_gross', 'net_pay', 'custom_formula', 'net_pay <= gross_pay', 'Net pay cannot exceed gross pay', True],
        ['taxes_range', 'taxes', 'range', '0:15000', 'Taxes must be between $0 and $15,000', True],
        ['taxes_positive', 'taxes', 'range', '0:', 'Taxes must be positive', True],
        ['taxes_less_than_gross', 'taxes', 'custom_formula', 'taxes <= gross_pay', 'Taxes cannot exceed gross pay', True],
        ['benefits_range', 'benefits', 'range', '0:10000', 'Benefits must be between $0 and $10,000', True],
        ['benefits_positive', 'benefits', 'range', '0:', 'Benefits must be positive', True],
        ['overtime_pay_range', 'overtime_pay', 'range', '0:10000', 'Overtime pay must be between $0 and $10,000', True],
        ['overtime_pay_positive', 'overtime_pay', 'range', '0:', 'Overtime pay must be positive', True],
        ['bonus_range', 'bonus', 'range', '0:20000', 'Bonus must be between $0 and $20,000', True],
        ['bonus_positive', 'bonus', 'range', '0:', 'Bonus must be positive', True]
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws_payroll.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data
    for row, rule in enumerate(payroll_rules, 2):
        for col, value in enumerate(rule, 1):
            cell = ws_payroll.cell(row=row, column=col, value=value)
            cell.border = border
    
    # Auto-adjust column widths
    for column in ws_payroll.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_payroll.column_dimensions[column_letter].width = adjusted_width
    
    # =============================================
    # DEPARTMENTS TABLE RULES
    # =============================================
    
    ws_departments = wb.create_sheet("Rules_departments")
    
    department_rules = [
        ['department_name_required', 'department_name', 'required', True, 'Department name is required', True],
        ['department_name_unique', 'department_name', 'unique', True, 'Department name must be unique', True],
        ['department_name_length', 'department_name', 'length', '1:100', 'Department name must be 1-100 characters', True],
        ['department_code_required', 'department_code', 'required', True, 'Department code is required', True],
        ['department_code_unique', 'department_code', 'unique', True, 'Department code must be unique', True],
        ['department_code_format', 'department_code', 'regex', r'^[A-Z]{2,10}$', 'Department code must be 2-10 uppercase letters', True],
        ['manager_id_exists', 'manager_id', 'foreign_key', 'employees.employee_id', 'Manager ID must exist in employees table', False],
        ['budget_amount_positive', 'budget_amount', 'range', '0:', 'Budget amount must be positive', False],
        ['budget_amount_range', 'budget_amount', 'range', '0:10000000', 'Budget amount must be between $0 and $10M', False]
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws_departments.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data
    for row, rule in enumerate(department_rules, 2):
        for col, value in enumerate(rule, 1):
            cell = ws_departments.cell(row=row, column=col, value=value)
            cell.border = border
    
    # Auto-adjust column widths
    for column in ws_departments.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_departments.column_dimensions[column_letter].width = adjusted_width
    
    # =============================================
    # PROJECTS TABLE RULES
    # =============================================
    
    ws_projects = wb.create_sheet("Rules_projects")
    
    project_rules = [
        ['project_code_required', 'project_code', 'required', True, 'Project code is required', True],
        ['project_code_unique', 'project_code', 'unique', True, 'Project code must be unique', True],
        ['project_code_format', 'project_code', 'regex', r'^PRJ\d{3}$', 'Project code must be in format PRJ001, PRJ002, etc.', True],
        ['project_name_required', 'project_name', 'required', True, 'Project name is required', True],
        ['project_name_length', 'project_name', 'length', '1:200', 'Project name must be 1-200 characters', True],
        ['client_name_length', 'client_name', 'length', '0:100', 'Client name must be 100 characters or less', False],
        ['start_date_format', 'start_date', 'data_type', 'date', 'Start date must be a valid date', False],
        ['end_date_format', 'end_date', 'data_type', 'date', 'End date must be a valid date', False],
        ['end_date_after_start', 'end_date', 'custom_formula', 'end_date > start_date', 'End date must be after start date', False],
        ['budget_positive', 'budget', 'range', '0:', 'Budget must be positive', False],
        ['budget_range', 'budget', 'range', '0:10000000', 'Budget must be between $0 and $10M', False],
        ['status_enum', 'status', 'enum', 'ACTIVE,COMPLETED,ON_HOLD,CANCELLED', 'Status must be one of: ACTIVE, COMPLETED, ON_HOLD, CANCELLED', True],
        ['status_default', 'status', 'default', 'ACTIVE', 'Status defaults to ACTIVE', True]
    ]
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws_projects.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data
    for row, rule in enumerate(project_rules, 2):
        for col, value in enumerate(rule, 1):
            cell = ws_projects.cell(row=row, column=col, value=value)
            cell.border = border
    
    # Auto-adjust column widths
    for column in ws_projects.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_projects.column_dimensions[column_letter].width = adjusted_width
    
    # =============================================
    # INSTRUCTIONS SHEET
    # =============================================
    
    ws_instructions = wb.create_sheet("Instructions")
    
    instructions_data = [
        ['Rule Types', 'Description', 'Example Value', 'Notes'],
        ['required', 'Field must not be null or empty', 'True', 'Use True for required fields'],
        ['range', 'Numeric value must be within range', 'min:max or min-max', 'Use colon or dash to separate min/max'],
        ['regex', 'String must match regular expression', '^[A-Z]{2}\\d{4}$', 'Standard regex patterns'],
        ['enum', 'Value must be in predefined list', 'value1,value2,value3', 'Comma-separated values'],
        ['data_type', 'Value must be of specific type', 'numeric, date, string, boolean', 'Supported data types'],
        ['custom_formula', 'Custom validation logic', 'date <= today()', 'Custom business logic'],
        ['unique', 'Value must be unique in the table', 'True', 'Ensures no duplicates'],
        ['foreign_key', 'Value must exist in another table', 'table.column', 'References another table'],
        ['length', 'String length must be within range', 'min:max', 'Character count validation'],
        ['default', 'Default value for the field', 'ACTIVE', 'Default value if not provided']
    ]
    
    # Add headers
    for col, header in enumerate(instructions_data[0], 1):
        cell = ws_instructions.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data
    for row, instruction in enumerate(instructions_data[1:], 2):
        for col, value in enumerate(instruction, 1):
            cell = ws_instructions.cell(row=row, column=col, value=value)
            cell.border = border
    
    # Auto-adjust column widths
    for column in ws_instructions.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_instructions.column_dimensions[column_letter].width = adjusted_width
    
    # =============================================
    # METADATA SHEET
    # =============================================
    
    ws_metadata = wb.create_sheet("Metadata")
    
    metadata_data = [
        ['Property', 'Value'],
        ['Created Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['Version', '1.0.0'],
        ['Description', 'Labor Analytics Validation Rules'],
        ['Total Tables', '5'],
        ['Total Rules', str(len(employee_rules) + len(time_rules) + len(payroll_rules) + len(department_rules) + len(project_rules))],
        ['Active Rules', str(sum(1 for rule in employee_rules + time_rules + payroll_rules + department_rules + project_rules if rule[5]))],
        ['Tables Covered', 'employees, time_tracking, payroll, departments, projects'],
        ['Rule Types', 'required, range, regex, enum, data_type, custom_formula, unique, foreign_key, length, default']
    ]
    
    # Add data
    for row, metadata in enumerate(metadata_data, 1):
        for col, value in enumerate(metadata, 1):
            cell = ws_metadata.cell(row=row, column=col, value=value)
            if row == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            cell.border = border
    
    # Auto-adjust column widths
    for column in ws_metadata.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_metadata.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    output_path = "validation_rules.xlsx"
    wb.save(output_path)
    
    print(f"✅ Excel validation rules file created: {output_path}")
    print(f"📊 Total rules created: {len(employee_rules) + len(time_rules) + len(payroll_rules) + len(department_rules) + len(project_rules)}")
    print(f"📋 Tables covered: employees, time_tracking, payroll, departments, projects")
    print(f"📖 Instructions sheet included for rule types and usage")


if __name__ == "__main__":
    create_validation_rules_excel()
