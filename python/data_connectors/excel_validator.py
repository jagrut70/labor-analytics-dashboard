"""
Excel Validation Rules Processor
Handles Excel-based validation rules for data quality assurance
"""

import pandas as pd
import openpyxl
import json
import re
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime, date
import numpy as np
import structlog
from pathlib import Path

logger = structlog.get_logger(__name__)


class ExcelValidator:
    """Excel-based validation rules processor"""
    
    def __init__(self, rules_file_path: str = "../excel/validation_rules.xlsx"):
        """Initialize Excel validator with rules file"""
        self.rules_file_path = rules_file_path
        self.rules = {}
        self._load_validation_rules()
    
    def _load_validation_rules(self):
        """Load validation rules from Excel file"""
        try:
            if not Path(self.rules_file_path).exists():
                logger.warning(f"Validation rules file not found: {self.rules_file_path}")
                self._create_default_rules()
                return
            
            # Load Excel workbook
            workbook = openpyxl.load_workbook(self.rules_file_path, data_only=True)
            
            # Process each worksheet
            for sheet_name in workbook.sheetnames:
                if sheet_name.startswith('Rules_'):
                    table_name = sheet_name.replace('Rules_', '')
                    self.rules[table_name] = self._parse_rules_sheet(workbook[sheet_name])
            
            logger.info(f"Loaded validation rules for {len(self.rules)} tables")
            
        except Exception as e:
            logger.error(f"Error loading validation rules: {e}")
            self._create_default_rules()
    
    def _parse_rules_sheet(self, worksheet) -> List[Dict[str, Any]]:
        """Parse rules from a worksheet"""
        rules = []
        
        # Find the data range
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row[0] or pd.isna(row[0]):  # Skip empty rows
                continue
            
            rule = {
                'rule_name': str(row[0]),
                'field_name': str(row[1]),
                'rule_type': str(row[2]),
                'rule_value': row[3],
                'error_message': str(row[4]) if row[4] else '',
                'is_active': bool(row[5]) if len(row) > 5 else True
            }
            
            # Parse rule value based on type
            rule['parsed_value'] = self._parse_rule_value(rule['rule_type'], rule['rule_value'])
            rules.append(rule)
        
        return rules
    
    def _parse_rule_value(self, rule_type: str, rule_value: Any) -> Any:
        """Parse rule value based on rule type"""
        if pd.isna(rule_value):
            return None
        
        try:
            if rule_type == 'range':
                # Parse range like "min:max" or "min-max"
                if isinstance(rule_value, str):
                    if ':' in rule_value:
                        min_val, max_val = rule_value.split(':')
                    elif '-' in rule_value:
                        min_val, max_val = rule_value.split('-')
                    else:
                        return rule_value
                    
                    return {
                        'min': float(min_val.strip()) if min_val.strip() != '' else None,
                        'max': float(max_val.strip()) if max_val.strip() != '' else None
                    }
            
            elif rule_type == 'regex':
                return str(rule_value)
            
            elif rule_type == 'enum':
                # Parse comma-separated values
                if isinstance(rule_value, str):
                    return [val.strip() for val in rule_value.split(',')]
                return rule_value
            
            elif rule_type == 'required':
                return bool(rule_value)
            
            elif rule_type == 'data_type':
                return str(rule_value)
            
            elif rule_type == 'custom_formula':
                return str(rule_value)
            
            else:
                return rule_value
                
        except Exception as e:
            logger.error(f"Error parsing rule value: {e}")
            return rule_value
    
    def _create_default_rules(self):
        """Create default validation rules if Excel file doesn't exist"""
        self.rules = {
            'employees': [
                {
                    'rule_name': 'employee_id_required',
                    'field_name': 'employee_id',
                    'rule_type': 'required',
                    'rule_value': True,
                    'error_message': 'Employee ID is required',
                    'is_active': True,
                    'parsed_value': True
                },
                {
                    'rule_name': 'salary_range',
                    'field_name': 'salary',
                    'rule_type': 'range',
                    'rule_value': '20000:200000',
                    'error_message': 'Salary must be between $20,000 and $200,000',
                    'is_active': True,
                    'parsed_value': {'min': 20000, 'max': 200000}
                },
                {
                    'rule_name': 'email_format',
                    'field_name': 'email',
                    'rule_type': 'regex',
                    'rule_value': r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$',
                    'error_message': 'Invalid email format',
                    'is_active': True,
                    'parsed_value': r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                }
            ],
            'time_tracking': [
                {
                    'rule_name': 'hours_worked_range',
                    'field_name': 'hours_worked',
                    'rule_type': 'range',
                    'rule_value': '0:24',
                    'error_message': 'Hours worked must be between 0 and 24',
                    'is_active': True,
                    'parsed_value': {'min': 0, 'max': 24}
                },
                {
                    'rule_name': 'date_not_future',
                    'field_name': 'date',
                    'rule_type': 'custom_formula',
                    'rule_value': 'date <= today()',
                    'error_message': 'Date cannot be in the future',
                    'is_active': True,
                    'parsed_value': 'date <= today()'
                }
            ]
        }
        logger.info("Created default validation rules")
    
    def validate_dataframe(self, df: pd.DataFrame, table_name: str) -> Dict[str, Any]:
        """Validate a DataFrame against Excel-based rules"""
        if table_name not in self.rules:
            logger.warning(f"No validation rules found for table: {table_name}")
            return {'valid': True, 'errors': [], 'warnings': []}
        
        validation_results = {
            'valid': True,
            'errors': [],
            'warnings': [],
            'total_records': len(df),
            'valid_records': len(df),
            'invalid_records': 0
        }
        
        table_rules = self.rules[table_name]
        
        for rule in table_rules:
            if not rule.get('is_active', True):
                continue
            
            field_name = rule['field_name']
            if field_name not in df.columns:
                logger.warning(f"Field {field_name} not found in DataFrame for table {table_name}")
                continue
            
            field_errors = self._validate_field(df, field_name, rule)
            validation_results['errors'].extend(field_errors)
            
            if field_errors:
                validation_results['valid'] = False
                validation_results['invalid_records'] = len(set([error['record_index'] for error in field_errors]))
                validation_results['valid_records'] = validation_results['total_records'] - validation_results['invalid_records']
        
        return validation_results
    
    def _validate_field(self, df: pd.DataFrame, field_name: str, rule: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Validate a specific field against a rule"""
        errors = []
        rule_type = rule['rule_type']
        parsed_value = rule['parsed_value']
        
        for index, value in df[field_name].items():
            try:
                if rule_type == 'required':
                    if pd.isna(value) or value == '':
                        errors.append({
                            'record_index': index,
                            'field_name': field_name,
                            'rule_name': rule['rule_name'],
                            'error_message': rule['error_message'],
                            'value': value
                        })
                
                elif rule_type == 'range':
                    if not pd.isna(value):
                        if parsed_value.get('min') is not None and value < parsed_value['min']:
                            errors.append({
                                'record_index': index,
                                'field_name': field_name,
                                'rule_name': rule['rule_name'],
                                'error_message': rule['error_message'],
                                'value': value
                            })
                        elif parsed_value.get('max') is not None and value > parsed_value['max']:
                            errors.append({
                                'record_index': index,
                                'field_name': field_name,
                                'rule_name': rule['rule_name'],
                                'error_message': rule['error_message'],
                                'value': value
                            })
                
                elif rule_type == 'regex':
                    if not pd.isna(value) and not re.match(parsed_value, str(value)):
                        errors.append({
                            'record_index': index,
                            'field_name': field_name,
                            'rule_name': rule['rule_name'],
                            'error_message': rule['error_message'],
                            'value': value
                        })
                
                elif rule_type == 'enum':
                    if not pd.isna(value) and value not in parsed_value:
                        errors.append({
                            'record_index': index,
                            'field_name': field_name,
                            'rule_name': rule['rule_name'],
                            'error_message': rule['error_message'],
                            'value': value
                        })
                
                elif rule_type == 'data_type':
                    if not pd.isna(value):
                        expected_type = parsed_value.lower()
                        if expected_type == 'numeric' and not isinstance(value, (int, float, np.number)):
                            errors.append({
                                'record_index': index,
                                'field_name': field_name,
                                'rule_name': rule['rule_name'],
                                'error_message': rule['error_message'],
                                'value': value
                            })
                        elif expected_type == 'date' and not isinstance(value, (date, datetime)):
                            errors.append({
                                'record_index': index,
                                'field_name': field_name,
                                'rule_name': rule['rule_name'],
                                'error_message': rule['error_message'],
                                'value': value
                            })
                
                elif rule_type == 'custom_formula':
                    # Handle custom formulas
                    if 'date <= today()' in parsed_value and field_name == 'date':
                        today = date.today()
                        if isinstance(value, str):
                            try:
                                value_date = pd.to_datetime(value).date()
                            except:
                                value_date = None
                        elif isinstance(value, (date, datetime)):
                            value_date = value.date() if isinstance(value, datetime) else value
                        else:
                            value_date = None
                        
                        if value_date and value_date > today:
                            errors.append({
                                'record_index': index,
                                'field_name': field_name,
                                'rule_name': rule['rule_name'],
                                'error_message': rule['error_message'],
                                'value': value
                            })
            
            except Exception as e:
                logger.error(f"Error validating field {field_name} with rule {rule['rule_name']}: {e}")
                errors.append({
                    'record_index': index,
                    'field_name': field_name,
                    'rule_name': rule['rule_name'],
                    'error_message': f"Validation error: {str(e)}",
                    'value': value
                })
        
        return errors
    
    def create_validation_report(self, validation_results: Dict[str, Any], output_path: str = None) -> str:
        """Create a detailed validation report"""
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"../excel/validation_report_{timestamp}.xlsx"
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = {
                'Metric': [
                    'Total Records',
                    'Valid Records',
                    'Invalid Records',
                    'Validation Rate (%)',
                    'Total Errors',
                    'Validation Status'
                ],
                'Value': [
                    validation_results['total_records'],
                    validation_results['valid_records'],
                    validation_results['invalid_records'],
                    round((validation_results['valid_records'] / validation_results['total_records']) * 100, 2) if validation_results['total_records'] > 0 else 0,
                    len(validation_results['errors']),
                    'PASS' if validation_results['valid'] else 'FAIL'
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Errors sheet
            if validation_results['errors']:
                errors_df = pd.DataFrame(validation_results['errors'])
                errors_df.to_excel(writer, sheet_name='Errors', index=False)
            
            # Rules sheet
            rules_data = []
            for table_name, rules in self.rules.items():
                for rule in rules:
                    rules_data.append({
                        'Table': table_name,
                        'Rule Name': rule['rule_name'],
                        'Field': rule['field_name'],
                        'Type': rule['rule_type'],
                        'Active': rule.get('is_active', True)
                    })
            
            if rules_data:
                rules_df = pd.DataFrame(rules_data)
                rules_df.to_excel(writer, sheet_name='Rules', index=False)
        
        logger.info(f"Validation report created: {output_path}")
        return output_path
    
    def export_rules_template(self, output_path: str = "../excel/validation_rules_template.xlsx"):
        """Export a template for validation rules"""
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Template for employees table
            employee_rules = [
                ['employee_id_required', 'employee_id', 'required', True, 'Employee ID is required', True],
                ['salary_range', 'salary', 'range', '20000:200000', 'Salary must be between $20,000 and $200,000', True],
                ['email_format', 'email', 'regex', r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', 'Invalid email format', True],
                ['hire_date_format', 'hire_date', 'data_type', 'date', 'Hire date must be a valid date', True]
            ]
            
            employee_df = pd.DataFrame(employee_rules, columns=[
                'Rule Name', 'Field Name', 'Rule Type', 'Rule Value', 'Error Message', 'Is Active'
            ])
            employee_df.to_excel(writer, sheet_name='Rules_employees', index=False)
            
            # Template for time_tracking table
            time_rules = [
                ['hours_worked_range', 'hours_worked', 'range', '0:24', 'Hours worked must be between 0 and 24', True],
                ['date_not_future', 'date', 'custom_formula', 'date <= today()', 'Date cannot be in the future', True],
                ['employee_id_required', 'employee_id', 'required', True, 'Employee ID is required', True]
            ]
            
            time_df = pd.DataFrame(time_rules, columns=[
                'Rule Name', 'Field Name', 'Rule Type', 'Rule Value', 'Error Message', 'Is Active'
            ])
            time_df.to_excel(writer, sheet_name='Rules_time_tracking', index=False)
            
            # Instructions sheet
            instructions = [
                ['Rule Types', 'Description', 'Example Value'],
                ['required', 'Field must not be null or empty', 'True'],
                ['range', 'Numeric value must be within range', 'min:max or min-max'],
                ['regex', 'String must match regular expression', '^[A-Z]{2}\\d{4}$'],
                ['enum', 'Value must be in predefined list', 'value1,value2,value3'],
                ['data_type', 'Value must be of specific type', 'numeric, date, string'],
                ['custom_formula', 'Custom validation logic', 'date <= today()']
            ]
            
            instructions_df = pd.DataFrame(instructions[1:], columns=instructions[0])
            instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
        
        logger.info(f"Validation rules template created: {output_path}")
        return output_path


if __name__ == "__main__":
    # Test the Excel validator
    validator = ExcelValidator()
    
    # Create sample data
    sample_data = pd.DataFrame({
        'employee_id': ['EMP001', 'EMP002', '', 'EMP004'],
        'salary': [75000, 85000, 15000, 250000],
        'email': ['john@company.com', 'jane@company.com', 'invalid-email', 'mike@company.com'],
        'hire_date': ['2020-01-15', '2019-03-20', '2021-06-10', '2020-11-05']
    })
    
    # Validate the data
    results = validator.validate_dataframe(sample_data, 'employees')
    
    print("Validation Results:")
    print(f"Valid: {results['valid']}")
    print(f"Total Records: {results['total_records']}")
    print(f"Valid Records: {results['valid_records']}")
    print(f"Invalid Records: {results['invalid_records']}")
    print(f"Total Errors: {len(results['errors'])}")
    
    if results['errors']:
        print("\nErrors:")
        for error in results['errors']:
            print(f"  Row {error['record_index']}: {error['field_name']} - {error['error_message']}")
    
    # Create validation report
    report_path = validator.create_validation_report(results)
    print(f"\nValidation report created: {report_path}")
